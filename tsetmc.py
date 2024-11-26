from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

import time
from bs4 import BeautifulSoup
import pandas as pd
import openpyxl
import re

from database import *
    

class tsetmc:
    def __init__(self):
        # Create an empty DataFrame for saving scraping tsetmc data
        self.columns=['نماد', 'نام', 
                'تعداد', 'حجم', 'ارزش', 'دیروز', 'اولین', 
                'مقدار آخرین معامله', 'تغییر آخرین معامله', 'درصد آخرین معامله',
                'مقدار قیمت پایانی', 'تغییر قیمت پایانی', 'درصد قیمت پایانی', 
                'کمترین', 'بیشترین', 
                'P/E', 'EPS', 
                'تعداد خرید', 'حجم خرید', 'قیمت خرید',
                'قیمت فروش', 'حجم فروش', 'تعداد فروش']

        self.df = pd.DataFrame(columns=self.columns)

        self.option_symbols = self.option_and_underlying_asset_symbols()


    # Function to convert the values to numbers
    @staticmethod
    def convert_to_number(value):
        # Strip any extra whitespace
        value = value.strip()
        
        # Remove commas from the number
        value = value.replace(',', '')
        
        # If the value ends with 'B', convert it to a billion
        if 'B' in value:
            return float(value.replace('B', '')) * 1_000_000_000
        
        # If the value ends with 'M', convert it to a million
        elif 'M' in value:
            return float(value.replace('M', '')) * 1_000_000
        
        # If ther is any '-' in tsetmc vlues it can be for negetive numbers or loss data
        elif '-' in value:
            if any(char.isdigit() for char in value):
                return float(value)
            else:
                return 0
            
        elif 'Infinity' in value:
            return 0
            
        # Otherwise, convert to integer or float as needed
        else:
            return float(value)


    @staticmethod
    def clear_excel_file(file_path, keep_headers=False):
        # Load the Excel file
        workbook = openpyxl.load_workbook(file_path)

        # Loop through all the sheets in the workbook
        for sheet in workbook.worksheets:
            # Determine the number of rows and columns in the sheet
            max_row = sheet.max_row
            max_column = sheet.max_column

            # If keeping headers, start from the second row, otherwise start from the first
            start_row = 2 if keep_headers else 1

            # Loop through all cells and clear their contents
            for row in range(start_row, max_row + 1):
                for col in range(1, max_column + 1):
                    sheet.cell(row=row, column=col).value = None

        # Save the modified workbook
        workbook.save(file_path)


    def save_option_data(self, df):
        # Create an empty DataFrame with the same columns as self.df
        options = pd.DataFrame(columns=self.columns)
        underlying_assets = pd.DataFrame(columns=self.columns)
        underlying_assets_list = []

        # Filter option rows
        for n in range(len(self.df)):
            if 'اختيار' in self.df['نام'][n]:
                # Save option rowes to options data frame
                options = options._append(self.df.loc[n], ignore_index=True)

        for n in range(len(options)):
            # Use regex to remove digits, hyphens, and the word related to call or put option
            cleaned_string = re.sub(r'\d+|-|اختيارخ|/|اختيارف', '', options['نام'][n], flags=re.IGNORECASE)
            # Return the cleaned string, also stripping any extra spaces that may result
            underlying_asset_symbol = cleaned_string.strip()
            if underlying_asset_symbol not in underlying_assets_list:
                underlying_assets_list.append(underlying_asset_symbol)

        for symbol in underlying_assets_list:
            for n in range(len(self.df)):
                if symbol == self.df['نماد'][n] :
                    # Save option rowes to options data frame
                    underlying_assets = underlying_assets._append(self.df.loc[n], ignore_index=True)

        # Save the DataFrame to database
        Database().save_dataframe(options, 'option_data')
        Database().save_dataframe(underlying_assets, 'underlying_assets')


    def option_and_underlying_asset_symbols(self):
        option_df = Database().load_table('option_data')
        underlying_asset_df = Database().load_table('underlying_assets')

        option_symbols = option_df['نماد'].tolist()
        underlying_asset_symbols = underlying_asset_df['نماد'].tolist()

        all = option_symbols + underlying_asset_symbols

        return all


    def update_all_data(self, just_option=False, repeat=-1):
        driver = webdriver.Firefox()
        driver.get('https://old.tsetmc.com/Loader.aspx?ParTree=15131F#')
        time.sleep(20)

        # Open the setting window
        setting_button = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, '//*[@id="id1"]'))
        )
        setting_button.click()

        
        # Wait for the modal to be visible
        modal = WebDriverWait(driver, 10).until(
            EC.visibility_of_element_located((By.ID, 'ModalWindowInner1'))  # Use the modal's ID
        )

        # Click on all symbol button
        all_button = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, "//div[@aria-label='نمایش همه نمادها در دیده بان']"))
        )
        all_button.click()



        while True:
            # If user enter repeat, the code will run by count of it
            if repeat == -1:
                pass
            elif repeat == 0:
                driver.quit()
                break
            else:
                repeat -= 1

            # Get page source
            html_content = driver.page_source

            # Parse HTML content with BeautifulSoup
            soup = BeautifulSoup(html_content, 'html.parser')  

            # Find all bourse elements with the specified class '{c}'
            elements = soup.find_all(class_="{c}")

            # Save each element in a list
            sorted_elements = [str(element) for element in elements]

            # Save each element information in a dictionary
            for element in sorted_elements:
                take_data_flag = True
            
                soup = BeautifulSoup(str(element), 'html.parser')

                # Find symbol and name of the element
                parameters = soup.find_all(class_="inst")

                # Save each element in a list
                sorted_parameters = [str(parameter) for parameter in parameters]

                # Parse the HTML string
                soup_symbol = BeautifulSoup(sorted_parameters[0], 'html.parser')
                symbol = soup_symbol.find('a').text

                soup_name = BeautifulSoup(sorted_parameters[1], 'html.parser')
                name = soup_name.find('a').text

                if just_option:
                    if symbol in self.option_symbols:
                        take_data_flag = True
                    else:
                        take_data_flag = False 


                if take_data_flag:
                    # Find element with 't0c5' class, some datas in list has this class
                    parameters = soup.find_all(class_="t0c5")

                    # Save each element in a list
                    sorted_parameters = [str(parameter) for parameter in parameters]

                    # Parse the HTML string
                    soup_value0 = BeautifulSoup(sorted_parameters[0], 'html.parser')
                    number = soup_value0.find('div').text
                    number = self.convert_to_number(number)

                    soup_value1 = BeautifulSoup(sorted_parameters[1], 'html.parser')
                    volum = soup_value1.find('div').text
                    volum = self.convert_to_number(volum)

                    soup_value2 = BeautifulSoup(sorted_parameters[2], 'html.parser')
                    value = soup_value2.find('div').text
                    value = self.convert_to_number(value)

                    soup_value3 = BeautifulSoup(sorted_parameters[3], 'html.parser')
                    last_day = soup_value3.find('div').text
                    last_day = self.convert_to_number(last_day)

                    soup_value4 = BeautifulSoup(sorted_parameters[4], 'html.parser')
                    first = soup_value4.find('div').text
                    first = self.convert_to_number(first)

                    # We have a jump in list columns
                    soup_value5 = BeautifulSoup(sorted_parameters[5], 'html.parser')
                    lowest = soup_value5.find('div').text
                    lowest = self.convert_to_number(lowest)

                    soup_value6 = BeautifulSoup(sorted_parameters[6], 'html.parser')
                    highest = soup_value6.find('div').text
                    highest = self.convert_to_number(highest)

                    soup_value7 = BeautifulSoup(sorted_parameters[7], 'html.parser')
                    eps = soup_value7.find('div').text
                    eps = self.convert_to_number(eps)

                    soup_value8 = BeautifulSoup(sorted_parameters[8], 'html.parser')
                    buyer_number = soup_value8.find('div').text
                    buyer_number = self.convert_to_number(buyer_number)

                    soup_value9 = BeautifulSoup(sorted_parameters[9], 'html.parser')
                    buyer_volum = soup_value9.find('div').text
                    buyer_volum = self.convert_to_number(buyer_volum)

                    soup_value10 = BeautifulSoup(sorted_parameters[10], 'html.parser')
                    seller_price = soup_value10.find('div').text
                    seller_price = self.convert_to_number(seller_price)

                    soup_value11 = BeautifulSoup(sorted_parameters[11], 'html.parser')
                    seller_volum = soup_value11.find('div').text
                    seller_volum = self.convert_to_number(seller_volum)



                    # Find element with 't0c' class, some datas in list has this class
                    parameters = soup.find_all(class_="t0c")

                    # Save each element in a list
                    sorted_parameters = [str(parameter) for parameter in parameters]

                    # Parse the HTML string
                    soup_value2 = BeautifulSoup(sorted_parameters[2], 'html.parser')
                    lasttrade_amount = soup_value2.find('div').text
                    lasttrade_amount = self.convert_to_number(lasttrade_amount)

                    soup_value3 = BeautifulSoup(sorted_parameters[3], 'html.parser')
                    lasttrade_change = soup_value3.find('div').text
                    lasttrade_change = self.convert_to_number(lasttrade_change)

                    soup_value4 = BeautifulSoup(sorted_parameters[4], 'html.parser')
                    lasttrade_percent = soup_value4.find('div').text
                    lasttrade_percent = self.convert_to_number(lasttrade_percent)

                    soup_value5 = BeautifulSoup(sorted_parameters[5], 'html.parser')
                    lastprice_amount = soup_value5.find('div').text
                    lastprice_amount = self.convert_to_number(lastprice_amount)

                    soup_value6 = BeautifulSoup(sorted_parameters[6], 'html.parser')
                    lastprice_change = soup_value6.find('div').text
                    lastprice_change = self.convert_to_number(lastprice_change)

                    soup_value7 = BeautifulSoup(sorted_parameters[7], 'html.parser')
                    lastprice_percent = soup_value7.find('div').text
                    lastprice_percent = self.convert_to_number(lastprice_percent)

                    soup_value8 = BeautifulSoup(sorted_parameters[8], 'html.parser')
                    pe = soup_value8.find('div').text
                    pe = self.convert_to_number(pe)

                    soup_value9 = BeautifulSoup(sorted_parameters[9], 'html.parser')
                    buyer_price = soup_value9.find('div').text
                    buyer_price = self.convert_to_number(buyer_price)

                    soup_value10 = BeautifulSoup(sorted_parameters[10], 'html.parser')
                    seller_number = soup_value10.find('div').text
                    seller_number = self.convert_to_number(seller_number)
                    

                    # Create a dictionary to save scraping data in  it.
                    dict = {
                        'نماد': symbol, 'نام': name, 'تعداد': number, 'حجم': volum, 'ارزش': value, 'دیروز': last_day, 'اولین': first, 
                        'مقدار آخرین معامله': lasttrade_amount, 'تغییر آخرین معامله': lasttrade_change, 'درصد آخرین معامله': lasttrade_percent, 
                        'مقدار قیمت پایانی': lastprice_amount, 'تغییر قیمت پایانی': lastprice_change, 'درصد قیمت پایانی': lastprice_percent, 
                        'کمترین': lowest, 'بیشترین': highest, 
                        'P/E': pe, 'EPS': eps,
                        'تعداد خرید': buyer_number, 'حجم خرید': buyer_volum, 'قیمت خرید': buyer_price, 
                        'قیمت فروش': seller_price, 'حجم فروش': seller_volum, 'تعداد فروش': seller_number
                    }

                    # Add the row to the DataFrame
                    self.df = self.df._append(dict, ignore_index=True)


            # Save the DataFrame to database
            Database().save_dataframe(self.df, 'tsetmc_data')

            # After take all update data, we update option data by it
            self.save_option_data(self.df)
            # Clear the DataFrame but keep the structure (i.e., column names)
            self.df = pd.DataFrame(columns=self.df.columns)

            time.sleep(1)

    
    

